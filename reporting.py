"""
reporting.py — HTML email generation, email dispatch, and Excel export.
"""
import os
import getpass
import smtplib
from datetime   import datetime
from email      import encoders
from email.mime.base      import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text      import MIMEText

import numpy  as np
import pandas as pd

from config import (
    SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD,
    REPORT_TO, ATR_MULTIPLIER, VERSION,
)
from utils import fv, fmt_sentiment, sentiment_color, signal_color

def generate_excel(df: pd.DataFrame, filepath: str) -> None:
    """Безопасный экспорт датафрейма в Excel."""
    if df.empty:
        print("Датафрейм пуст, экспорт отменен.")
        return
        
    # Создание директории, если она не существует
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    try:
        # Требуется установленная библиотека openpyxl (pip install openpyxl)
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Quant_Analysis')
    except Exception as e:
        print(f"Системная ошибка при записи Excel: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# HTML EMAIL
# ─────────────────────────────────────────────────────────────────────────────
def _row_html(r: pd.Series) -> str:
    sc    = sentiment_color(r["AI_Sentiment"])
    sgc   = signal_color(r["Signal"])
    badge = (
        f'<span style="background:#FFD700;color:#000;padding:1px 6px;'
        f'border-radius:3px;font-weight:bold;font-size:10px;">'
        f'{r["TOP_PICK"]}</span>'
    ) if r["TOP_PICK"] else ""
    bt_v = r["bt_pnl_pct"]
    bt_c = "#00C851" if (bt_v is not None and float(bt_v) >= 0) else "#D50000"
    bt_t = f"{float(bt_v):+.1f}%" if bt_v is not None else "—"
    sw   = " ⚠" if r["bt_stop_hit"] else ""
    return f"""<tr style="border-bottom:1px solid #2a2a3a;">
      <td style="padding:7px 10px;font-weight:600;">{r['Asset']}
        <br><span style="color:#888;font-size:11px;">{r['Ticker']} · {r['Currency']}</span>
        <br>{badge}</td>
      <td style="text-align:center;">
        <span style="background:{sgc};color:#fff;padding:2px 9px;border-radius:3px;
          font-weight:bold;font-size:12px;">{r['Signal']}</span></td>
      <td style="text-align:center;font-size:18px;font-weight:bold;">{r['Score'] or '—'}</td>
      <td style="text-align:center;color:{sc};font-weight:bold;">
        {fmt_sentiment(r['AI_Sentiment'])}<br>
        <span style="color:#777;font-size:10px;font-weight:normal;">
          {str(r['AI_Summary'])[:45]}</span></td>
      <td style="text-align:right;font-family:monospace;">{r['Entry_Fmt']}</td>
      <td style="text-align:right;font-family:monospace;color:#00C851;">{r['Target_Fmt']}</td>
      <td style="text-align:right;font-family:monospace;color:#FF6D00;">{r['Stop_Fmt']}</td>
      <td style="text-align:right;">
        {fv(r['Upside_Pct'], 1, '%', '+') if r['Upside_Pct'] is not None else '—'}</td>
      <td style="text-align:right;color:{bt_c};font-weight:bold;">{bt_t}{sw}</td>
      <td style="text-align:center;font-size:11px;color:#aaa;">{r['Geo_Risk']}</td>
    </tr>"""


def build_html_email(df: pd.DataFrame) -> str:
    top3      = df.nlargest(3, "Score")
    all_rows  = "".join(
        _row_html(r)
        for _, r in df.sort_values("Score", ascending=False).iterrows()
    )
    top3_html = ""
    for rank, (_, r) in enumerate(top3.iterrows(), 1):
        up  = f"{float(r['Upside_Pct']):+.1f}%" if r["Upside_Pct"] is not None else "—"
        bt  = f"{float(r['bt_pnl_pct']):+.1f}%" if r["bt_pnl_pct"] is not None else "—"
        bt_c = "#00C851" if r["bt_pnl_pct"] is not None and float(r["bt_pnl_pct"]) >= 0 else "#D50000"
        top3_html += f"""
        <div style="background:#16213e;border-left:4px solid #FFD700;padding:14px 18px;
                    border-radius:6px;margin-bottom:10px;">
          <div style="font-size:11px;color:#FFD700;font-weight:bold;">#{rank} TOP PICK</div>
          <div style="font-size:18px;font-weight:bold;color:#fff;margin:4px 0;">
            {r['Asset']}
            <span style="color:#888;font-size:13px;">({r['Ticker']} · {r['Currency']})</span>
          </div>
          <div style="display:flex;flex-wrap:wrap;gap:14px;font-size:12px;color:#aaa;">
            <span>Score: <b style="color:#fff">{r['Score']}/100</b></span>
            <span>Signal: <b style="color:{signal_color(r['Signal'])}">{r['Signal']}</b></span>
            <span>Entry: <b style="color:#fff">{r['Entry_Fmt']}</b></span>
            <span>Target: <b style="color:#00C851">{r['Target_Fmt']}</b></span>
            <span>Stop: <b style="color:#FF6D00">{r['Stop_Fmt']}</b></span>
            <span>Upside: <b style="color:#fff">{up}</b></span>
            <span>AI Sentiment:
              <b style="color:{sentiment_color(r['AI_Sentiment'])}">
                {fmt_sentiment(r['AI_Sentiment'])}</b></span>
            <span>12m BT: <b style="color:{bt_c}">{bt}</b></span>
          </div>
          <div style="margin-top:5px;color:#777;font-size:11px;font-style:italic;">
            {r['AI_Summary']}</div>
        </div>"""

    return f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
  body{{font-family:'Segoe UI',Arial,sans-serif;background:#0a0a1a;color:#ddd;margin:0;padding:0}}
  .wrap{{max-width:1100px;margin:0 auto;padding:20px}}
  h1{{color:#FFD700;font-size:22px;margin-bottom:2px}}
  h2{{color:#888;font-size:14px;font-weight:normal;margin-bottom:20px}}
  h3{{color:#FFD700;font-size:14px;margin:20px 0 10px;border-bottom:1px solid #222;padding-bottom:5px}}
  table{{width:100%;border-collapse:collapse;font-size:12px}}
  th{{background:#12122a;color:#FFD700;padding:9px 8px;text-align:left;font-size:11px;
      border-bottom:2px solid #FFD700;white-space:nowrap}}
  tr:hover{{background:#12122a}}
  .foot{{margin-top:20px;color:#444;font-size:10px;text-align:center}}
</style></head><body><div class="wrap">
  <h1>Trade Republic — Quant-AI Report v{VERSION}</h1>
  <h2>{datetime.now().strftime('%Y-%m-%d %H:%M')} &nbsp;|&nbsp; {len(df)} instruments
     &nbsp;|&nbsp; {_GEMINI_MODEL_DISPLAY} &nbsp;|&nbsp;
     Stop-Loss = {ATR_MULTIPLIER}×ATR</h2>
  <h3>TOP 3 PICKS</h3>{top3_html}
  <h3>Full Watchlist</h3>
  <table>
    <tr>
      <th>Asset / Ticker</th><th>Signal</th><th>Score</th>
      <th>AI Sentiment</th><th>Entry</th><th>Target</th>
      <th>Stop ({ATR_MULTIPLIER}×ATR)</th><th>Upside%</th>
      <th>12m Backtest</th><th>Geo</th>
    </tr>{all_rows}
  </table>
  <div class="foot">
    Auto-generated · {_GEMINI_MODEL_DISPLAY} sentiment · Not financial advice.
  </div>
</div></body></html>"""


# Import model name for display — done at module level to avoid circular imports.
try:
    from config import GEMINI_MODEL as _GEMINI_MODEL_DISPLAY
except ImportError:
    _GEMINI_MODEL_DISPLAY = "Gemini"


# ─────────────────────────────────────────────────────────────────────────────
# EMAIL SENDER
# ─────────────────────────────────────────────────────────────────────────────
def send_email_report(df: pd.DataFrame, csv_path: str, xlsx_path: str):
    """Send the HTML report with CSV and XLSX attachments via Gmail SMTP."""
    print("\n─── Email Report ────────────────────────────────────────────")
    sender    = SMTP_USER     or input("Sender Gmail:    ").strip()
    password  = SMTP_PASSWORD or getpass.getpass("App Password:    ")
    recipient = REPORT_TO     or input("Recipient email: ").strip()
    if not sender or not recipient:
        print("Skipping — credentials incomplete.")
        return

    msg            = MIMEMultipart("alternative")
    msg["Subject"] = f"Quant-AI v{VERSION} — {datetime.now().strftime('%Y-%m-%d')}"
    msg["From"]    = sender
    msg["To"]      = recipient
    msg.attach(MIMEText(build_html_email(df), "html"))

    tag = f"watchlist_v{VERSION.replace('.', '')}"
    for fpath, fname in [(csv_path, f"{tag}.csv"), (xlsx_path, f"{tag}.xlsx")]:
        if os.path.exists(fpath):
            with open(fpath, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition",
                                f'attachment; filename="{fname}"')
                msg.attach(part)
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as srv:
            srv.ehlo()
            srv.starttls()
            srv.login(sender, password)
            srv.sendmail(sender, recipient, msg.as_string())
        print(f"Report sent to {recipient}")
    except Exception as e:
        print(f"Email failed: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────
def write_excel(df: pd.DataFrame, path: str):
    """Write the results DataFrame to a formatted .xlsx file."""
    try:
        from openpyxl.styles import PatternFill, Font, Border, Side
        from openpyxl.utils  import get_column_letter

        FILLS  = {
            "BUY":  PatternFill("solid", fgColor="C6EFCE"),
            "SELL": PatternFill("solid", fgColor="FFC7CE"),
            "HOLD": PatternFill("solid", fgColor="FFEB9C"),
        }
        GOLD   = PatternFill("solid", fgColor="FFD700")
        BOLD   = Font(bold=True)
        thin   = Side(style="thin")
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

        want = [
            "Asset", "Ticker", "Currency", "Score",
            "Score_Fundamental", "Score_Technical", "Score_GeoSentiment",
            "Signal", "TOP_PICK", "AI_Sentiment", "AI_Summary",
            "Entry_Fmt", "Target_Fmt", "Stop_Fmt",
            "Entry_Price", "Target_Price", "Stop_Loss", "ATR",
            "Upside_Pct", "RSI", "Geo_Risk", "Geo_Score",
            "PE", "PEG", "ROE", "PB", "DivYield_pct",
            "CAGR_5Y_pct", "Vol_Ann_pct",
            "Corr_CPI", "Corr_CB_Rate", "Corr_Oil", "Corr_Gold", "Corr_SP500",
            "bt_signal", "bt_price_entry", "bt_price_now",
            "bt_pnl_pct", "bt_stop_hit", "bt_note", "Note",
        ]
        cols       = [c for c in want if c in df.columns]
        sheet_name = f"v{VERSION}"

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df[cols].to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"

            sig_col = cols.index("Signal")   + 1
            top_col = cols.index("TOP_PICK") + 1

            for col in ws.columns:
                w = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[
                    get_column_letter(col[0].column)
                ].width = min(w + 3, 48)

            for row in ws.iter_rows(min_row=2):
                sig = str(row[sig_col - 1].value or "")
                top = str(row[top_col - 1].value or "")
                for cell in row:
                    cell.border = BORDER
                row[sig_col - 1].fill = FILLS.get(sig, PatternFill())
                if "TOP PICK" in top:
                    for cell in row:
                        cell.fill = GOLD
                        cell.font = BOLD

        print(f"Excel saved: {path}")
    except Exception as e:
        print(f"Excel error: {e}")
